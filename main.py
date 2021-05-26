from tabula import read_pdf as rp
from pandas import DataFrame
from pandas import option_context
# from shutil import copy
from openpyxl import Workbook
from openpyxl import load_workbook
import traceback as tb
import time
import pprint as pp
from math import isnan
from tabulate import tabulate
from re import compile, DOTALL, MULTILINE


verbose = True
def main():
    df_dict = pdf_to_dict()
    # print(df_dict)

    try:
        # Parse date from dict:
        d = date(df_dict)
        # print_pretty_dict(df_dict)

        # expensive and overkill but makes life easier
        # I did this to split emps by 'Days', 'Mid Shift', 'Nights'
        rns = emp_by_shift(df_dict, 'RN')
        edts = emp_by_shift(df_dict, 'EDT')
        us = emp_by_shift(df_dict, 'US')

        rn_days, rn_mids, rn_nights = get_shift_type(rns)
        tech_days, tech_mids, tech_nights = get_shift_type(edts)
        sec_days, sec_mids, sec_nights = get_shift_type(us)
        wb = load_workbook(filename='template.xlsx')
        ws = wb.active
        days_chg = input('Enter who you want to be the Days CHARGE:')
    except Exception:
        tb.print_exc()
        time.sleep(0.1)
        print("Error in parsing data from dict")

    # Grab name to save new excel file as:
    # xlsx_name = str(input("What would you like to name the excel file?\n"))

    # Create workbook from template:
    wb = load_workbook(filename='template.xlsx')
    ws = wb.active


def get_shift_type(ppl):
    days = []
    mids = []
    nights = []
    for n in range(0, len(ppl)):
        if ppl[n][2] == 'Days':
            days += ppl[n]
        elif ppl[n][2] == 'Mid Shift':
            mids += ppl[n]
        else:
            nights += ppl[n]
    return days, mids, nights


# calling this over and over is a pretty expensive operation.
# I just didn't want to return a massive tuple of lists or anything
def emp_by_shift(df_dict, job):
    lst = []
    cols = list(x for x in df_dict[0])

    for i in range(len(df_dict)):
        e = str(df_dict[i][cols[0]])
        j = str(df_dict[i][cols[1]])
        s = str(df_dict[i][cols[2]])

        if e != 'nan' and 'Open Shift' not in e and 'TIMEOFF' not in s and job != 'ORIENTATION':

            shift = shift_to_list(s)
            shift_type = ''

            try:
                if 'P' in shift[0] and 'A' in shift[1]:
                    shift_type = 'Nights'
                elif 'A' in shift[0] and 'P' in shift[1]:
                    # not sure why regular str comparison wasn't working here
                    if int(shift[1].replace(':', '').replace('P', '')) > 830:
                        shift_type = 'Mid Shift'
                    else:
                        shift_type = 'Days'
                if job in j:
                    lst.append(tuple((e, shift, shift_type)))
                elif 'CHG' in j:
                    lst.append(tuple((e, shift, shift_type, 'CHG')))

            except Exception:
                # tb.print_exc()
                # print("Error parsing: " + str(s))
                continue

    return lst


def shift_to_list(s):
    shift = []
    for k in s.split():
        if ':' in k:
            if 'A' in k or 'P' in k:
                shift += [k]

    return shift


def date(d):
    assert isinstance(d, dict)
    return list(d[0].values())[2]


# def print_pretty_dict(df_dict):
#     assert isinstance(df_dict, dict)
#     cols = list(x for x in df_dict[0])
#     # all rows
#     for i in range(0, len(df_dict)):
#         if i == 0:
#             print(str.join('            ', (i for i in cols)))
#         else:
#             employee = df_dict[i][cols[0]]
#             job = df_dict[i][cols[1]]
#             shift = df_dict[i][cols[2]]
#             print(str.join('            ', [repr(employee), repr(job), repr(shift)]))


def pdf_to_dict():
    while True:
        try:
            pdf_name = str(input("Enter full path of the PDF to read: "))
            # Convert pdf info to DataFrame, then to dict
            first_df = DataFrame(rp(pdf_name, pages='all')[0])

            def count_pages(file_path):
                rxcountpages = compile(r"/Type\s*/Page([^s]|$)", MULTILINE | DOTALL)
                with open(file_path, "rb") as temp_file:
                    return len(rxcountpages.findall(str(temp_file.read())))

            pdf_len = count_pages(pdf_name)
            for pg in range(1, pdf_len):
                df2 = DataFrame(rp(pdf_name, pages='all')[pg])
                for i in range(0, len(df2)):
                    df2.rename(index={i: i + len(first_df)}, inplace=True)
                first_df = first_df.append(df2)

            print(first_df)
            print(tabulate(first_df, headers='keys', tablefmt='psql'))
            # if verbose:
            #     print(df, '\n')

            df_dict = first_df.to_dict('index')

            break
        except Exception:
            tb.print_exc()
            print("Error parsing PDF, please try again.")

    return df_dict


if __name__ == '__main__':
    main()